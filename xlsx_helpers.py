"""Общие помощники XLSX-экспорта.

Модуль не содержит логики конкретных листов; здесь находятся только
переиспользуемые расчёты и оформление worksheet.
"""

from __future__ import annotations

import logging
import math
from collections.abc import Mapping, Sequence
from functools import lru_cache
from typing import Any

from .models import RouteData
from .support import type_label

logger = logging.getLogger("wikiroutes.xlsx.helpers")


@lru_cache(maxsize=128)
def cached_type_label(route_type: object) -> str:
    """Кэширует преобразование типа транспорта в русскую метку."""
    return type_label(route_type)


def route_fully_in_bbox(
    route: RouteData,
    bbox: tuple[float, float, float, float] | None,
) -> bool:
    """Проверяет, целиком ли геометрия маршрута находится внутри bbox."""
    if bbox is None or not route.directions:
        return False

    min_lat, min_lon, max_lat, max_lon = bbox

    for direction in route.directions:
        for lat, lon in direction.coords:
            if not (
                math.isfinite(lat)
                and math.isfinite(lon)
                and min_lat <= lat <= max_lat
                and min_lon <= lon <= max_lon
            ):
                return False

    return True


def network_density(
    net_metrics: Mapping[str, Any] | None,
    bbox: tuple[float, float, float, float] | None,
) -> Any:
    """Плотность уникальной сети, км/км², по bbox города."""
    if not net_metrics or not bbox:
        return ""

    min_lat, min_lon, max_lat, max_lon = bbox
    lat_c = (min_lat + max_lat) / 2
    area_km2 = ((max_lat - min_lat) * 111.0) * (
        (max_lon - min_lon) * 111.0 * math.cos(math.radians(lat_c))
    )

    if area_km2 <= 0:
        return ""

    return round(net_metrics.get("unique_km", 0.0) / area_km2, 3)


def finish_sheet(
    ws: Any,
    headers: Sequence[object],
    num_fmt: Mapping[int, str] | None = None,
    wrap: set[int] | None = None,
    link_col: int | None = None,
    fixed_widths: Sequence[int] | None = None,
    inline_styled: bool = False,
) -> None:
    """Оформляет worksheet: шапка, ширины, форматы, ссылки и перенос."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    head_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    head_font = Font(color="FFFFFF", bold=True)
    link_font = Font(color="0563C1", underline="single")

    for c in ws[1]:
        c.fill, c.font = head_fill, head_font
        c.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    ws.freeze_panes = "A2"

    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions

    if fixed_widths:
        for i, width in enumerate(fixed_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
    else:
        max_rows = min(ws.max_row, 200)
        for i, header in enumerate(headers, 1):
            width = len(str(header))
            for row in ws.iter_rows(
                min_row=2,
                max_row=max_rows,
                min_col=i,
                max_col=i,
            ):
                value = row[0].value
                if value is not None:
                    width = max(
                        width,
                        min(max(len(str(value)) for s in str(value).split("\n")), 60),
                    )
            ws.column_dimensions[get_column_letter(i)].width = min(width + 2, 62)

    if num_fmt and not inline_styled:
        for col, fmt in num_fmt.items():
            for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
                row[0].number_format = fmt

    if wrap and not inline_styled:
        for col in wrap:
            for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
                row[0].alignment = Alignment(wrap_text=True, vertical="top")

    if link_col and not inline_styled:
        for row in ws.iter_rows(min_row=2, min_col=link_col, max_col=link_col):
            if row[0].value:
                row[0].hyperlink = str(row[0].value)
                row[0].font = link_font


__all__ = ["cached_type_label", "finish_sheet", "network_density", "route_fully_in_bbox"]
