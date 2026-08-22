"""XLSX-экспорт сети маршрутов и связанных аналитических данных.

Каждый лист отчёта собирается отдельной функцией ``_write_*_sheet``;
общее оформление (шапка, ширины, форматы, ссылки) выполняет
``_finish_sheet``. ``build_xlsx`` только группирует данные и вызывает
построители в фиксированном порядке листов.
"""

from __future__ import annotations

import logging
import math
from collections.abc import Mapping, Sequence
from datetime import datetime
from functools import lru_cache
from pathlib import Path
from typing import Any

from .models import RouteData
from .support import (
    collect_terminal_routes,
    direction_endpoints,
    fix_stop_coord,
    fmt_curv_cell,
    type_label,
)

logger = logging.getLogger("wikiroutes.xlsx")


@lru_cache(maxsize=128)
def _cached_type_label(route_type: object) -> str:
    """Кэширует преобразование типа транспорта в русскую метку."""
    return type_label(route_type)


def _route_fully_in_bbox(
    route: RouteData, bbox: tuple[float, float, float, float] | None
) -> bool:
    """Проверяет, целиком ли геометрия маршрута находится внутри bbox."""
    if bbox is None or not route.directions:
        return False

    min_lat, min_lon, max_lat, max_lon = bbox

    for direction in route.directions:
        for lat, lon in direction.coords:
            if not (
                math.isfinite(lat)
                and math.isfinite(lon)
                and min_lat <= lat <= max_lat
                and min_lon <= lon <= max_lon
            ):
                return False

    return True


def _network_density(
    net_metrics: Mapping[str, Any] | None,
    bbox: tuple[float, float, float, float] | None,
) -> Any:
    """Плотность уникальной сети, км/км², по bbox города."""
    if not net_metrics or not bbox:
        return ""

    min_lat, min_lon, max_lat, max_lon = bbox

    lat_c = (min_lat + max_lat) / 2

    area_km2 = ((max_lat - min_lat) * 111.0) * (
        (max_lon - min_lon) * 111.0 * math.cos(math.radians(lat_c))
    )

    if area_km2 <= 0:
        return ""

    return round(net_metrics.get("unique_km", 0.0) / area_km2, 3)


# ═══════════════════════════════════════════════════════════════════════
# ОБЩЕЕ ОФОРМЛЕНИЕ ЛИСТА
# ═══════════════════════════════════════════════════════════════════════
def _finish_sheet(
    ws: Any,
    headers: Sequence[object],
    num_fmt: Mapping[int, str] | None = None,
    wrap: set[int] | None = None,
    link_col: int | None = None,
    fixed_widths: Sequence[int] | None = None,
    inline_styled: bool = False,
) -> None:
    """Шапка, закрепление, автофильтр, ширины, форматы, ссылки, перенос."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    head_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    head_font = Font(color="FFFFFF", bold=True)
    link_font = Font(color="0563C1", underline="single")

    # Стили для заголовков
    for c in ws[1]:
        c.fill, c.font = head_fill, head_font
        c.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    ws.freeze_panes = "A2"

    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions

    # Установка ширины колонок
    if fixed_widths:
        for i, w in enumerate(fixed_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    else:
        # Вычисляем ширину по первым 200 строкам (быстрее)
        max_rows = min(ws.max_row, 200)

        for i, h in enumerate(headers, 1):
            w = len(str(h))

            for row in ws.iter_rows(
                min_row=2,
                max_row=max_rows,
                min_col=i,
                max_col=i,
            ):
                v = row[0].value

                if v is not None:
                    # Берём максимальную длину строки (до 60 символов)
                    w = max(
                        w,
                        min(max(len(str(v)) for s in str(v).split("\n")), 60),
                    )

            ws.column_dimensions[get_column_letter(i)].width = min(w + 2, 62)

    # Числовые форматы
    if num_fmt and not inline_styled:
        for col, fmt in num_fmt.items():
            for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
                row[0].number_format = fmt

    # Перенос текста
    if wrap and not inline_styled:
        for col in wrap:
            for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
                row[0].alignment = Alignment(wrap_text=True, vertical="top")

    # Гиперссылки
    if link_col and not inline_styled:
        for row in ws.iter_rows(min_row=2, min_col=link_col, max_col=link_col):
            if row[0].value:
                row[0].hyperlink = str(row[0].value)
                row[0].font = link_font


# ═══════════════════════════════════════════════════════════════════════
# ЛИСТ «СВОДКА»
# ═══════════════════════════════════════════════════════════════════════
def _summary_type_rows(
    ok: list[RouteData],
    bbox: tuple[float, float, float, float] | None,
) -> list[list[Any]]:
    """Строки сводки с количеством маршрутов по типам транспорта.

    Особенность оригинала: поезда считаются только целиком внутри bbox,
    остальные типы — без географического фильтра.
    """
    rows: list[list[Any]] = [
        [
            title,
            sum(1 for r in ok if _cached_type_label(r.route_type) == label),
        ]
        for title, label in (
            ("Троллейбусов", "Троллейбус"),
            ("Трамваев", "Трамвай"),
            ("Автобусов", "Автобус"),
            ("Маршруток", "Маршрутное такси"),
            ("Водных маршрутов", "Водный транспорт"),
        )
    ]

    # В исходном отчёте «Поездов» стоит перед остальными типами
    trains_in_bbox = sum(
        1
        for r in ok
        if _cached_type_label(r.route_type) == "Поезд" and _route_fully_in_bbox(r, bbox)
    )

    return [["Поездов", trains_in_bbox], *rows]


def _write_summary_sheet(wb: Any, rows: Sequence[Sequence[Any]]) -> None:
    """Лист «Сводка»: пары параметр-значение."""
    ws = wb.active
    assert ws is not None
    ws.title = "Сводка"
    ws.append(["Параметр", "Значение"])

    for row in rows:
        ws.append(list(row))

    _finish_sheet(ws, ["Параметр", "Значение"], wrap={2}, fixed_widths=[25, 35])


def _summary_rows(
    *,
    city: str,
    city_title: str,
    routes: Sequence[RouteData],
    ok: list[RouteData],
    bad: list[RouteData],
    ideas: list[RouteData],
    skipped_inactive: int,
    curv_limit: float,
    minlen_limit: float,
    radius_limit: float,
    cut_curv: int,
    cut_len: int,
    cut_radius: int,
    dedup_removed: Sequence[Mapping[str, Any]] | None,
    dedup_analysis: Mapping[str, Any] | None,
    net_metrics: Mapping[str, Any] | None,
    net_density: Any,
    kml_routes: Sequence[RouteData] | None,
    unique_stops: Mapping[str, Mapping[str, Any]] | None,
    generated_routes: Sequence[Mapping[str, Any]] | None,
    bbox: tuple[float, float, float, float] | None,
    ghs_stats: Mapping[Any, Any] | None,
    ghs_meta: Mapping[str, Any] | None,
    poi_buffer_m: float,
    poi_stats: Mapping[Any, Any] | None,
    built_s_stats: Mapping[Any, Any] | None,
    built_s_meta: Mapping[str, Any] | None,
    overture_stats: Mapping[Any, Any] | None,
    overture_meta: Mapping[str, Any] | None,
) -> list[list[Any]]:
    """Все строки листа «Сводка»."""
    rows: list[list[Any]] = [
        ["Город", city_title],
        ["Slug", city],
        ["Дата", datetime.now().astimezone().strftime("%Y-%m-%d %H:%M")],
        ["Маршрутов в обработке", len(routes)],
        ["Успешно (в отчёте)", len(ok)],
        ["Ошибок", len(bad)],
        ["Идей пассажиров", len(ideas)],
        ["Пропущено неактивных (--active-only)", skipped_inactive],
        ["Отсечено: криволинейность", cut_curv],
        ["Отсечено: длина", cut_len],
        ["Отсечено: радиус", cut_radius],
        [
            "Дедупликация: удалено направлений",
            len(dedup_removed) if dedup_removed else 0,
        ],
        [
            "Дедупликация: Км до",
            dedup_analysis.get("km_coef", "") if dedup_analysis else "",
        ],
        [
            "Маршрутный коэффициент Км",
            net_metrics.get("km_coef", "") if net_metrics else "",
        ],
        [
            "Суммарная длина трасс, км",
            net_metrics.get("total_km", "") if net_metrics else "",
        ],
        [
            "Длина уникальной сети, км",
            net_metrics.get("unique_km", "") if net_metrics else "",
        ],
        ["Плотность сети, км/км²", net_density],
        ["Макс. коэфф. непрямолинейности (0=без лимита)", curv_limit],
        ["Мин. длина маршрута, км (0=без лимита)", minlen_limit],
        ["Радиус от центра, км (0=без лимита)", radius_limit],
        ["Маршрутов на карте (KML)", len(kml_routes) if kml_routes is not None else 0],
        ["Остановок уникальных", len(unique_stops) if unique_stops is not None else 0],
        ["Конечных", len(collect_terminal_routes(ok))],
        [
            "Сгенерировано маршрутов (Якимов)",
            len(generated_routes) if generated_routes else "",
        ],
    ]

    rows.extend(_summary_type_rows(ok, bbox))

    rows.extend(
        [
            ["GHS: буфер, м", ghs_meta.get("buffer_m", "") if ghs_meta else ""],
            [
                "GHS: суммарный объём, тыс. м³",
                round(sum(s.volume_m3 for s in ghs_stats.values()) / 1e3, 1)
                if ghs_stats
                else "",
            ],
            ["POI: буфер, м", poi_buffer_m if poi_stats else ""],
            [
                "POI: суммарное value",
                round(sum(st.total_value for st in poi_stats.values()), 1)
                if poi_stats
                else "",
            ],
        ]
    )

    if built_s_meta:
        rows.append(["GHS-BUILT-S: буфер, м", built_s_meta.get("buffer_m", "")])

    if built_s_stats:
        tot_s = sum(s.surface_m2 for s in built_s_stats.values())
        rows.append(
            ["GHS-BUILT-S: суммарная поверхность, тыс. м²", round(tot_s / 1e3, 1)]
        )

    if overture_meta:
        rows.append(["Overture: буфер, м", overture_meta.get("buffer_m", "")])

    if overture_stats:
        tot_a = sum(s.total_area_m2 for s in overture_stats.values())
        rows.append(["Overture: суммарная площадь, тыс. м²", round(tot_a / 1e3, 1)])

    return rows


# ═══════════════════════════════════════════════════════════════════════
# ЛИСТ «МАРШРУТЫ»
# ═══════════════════════════════════════════════════════════════════════
_ROUTES_BASE_HEADERS = [
    "Город",
    "Тип",
    "Номер",
    "ID",
    "Направлений",
    "Источник",
    "№ напр.",
    "Направление",
    "От",
    "До",
    "Длина, км",
    "Коэфф. непрямолинейности",
    "Остановок",
    "Ссылка",
]

_ROUTES_NUM_FMT_BY_HEADER = {
    "Длина, км": "0.0",
    "Коэфф. непрямолинейности": "0.00",
    "Остановок": "0",
    "Объём застройки, тыс. м³": "#,##0.0",
    "Площадь коридора, км²": "0.000",
    "Поверхность застройки, тыс. м²": "#,##0.0",
    "Площадь объектов, тыс. м²": "#,##0.0",
    "Точек POI": "0",
    "Суммарное value": "#,##0.0",
}

_ROUTES_FIXED_WIDTHS = [
    12,  # Город
    10,  # Тип
    12,  # Номер
    8,  # ID
    10,  # Направлений
    12,  # Источник
    8,  # № напр.
    20,  # Направление
    15,  # От
    15,  # До
    10,  # Длина, км
    10,  # Коэфф. непрямолинейности
    10,  # Остановок
    30,  # Ссылка
]


def _routes_headers(
    *,
    ghs_stats: Mapping[Any, Any] | None,
    built_s_stats: Mapping[Any, Any] | None,
    overture_stats: Mapping[Any, Any] | None,
    poi_stats: Mapping[Any, Any] | None,
) -> list[str]:
    """Заголовки листа «Маршруты» с опциональными аналитическими колонками."""
    headers = list(_ROUTES_BASE_HEADERS)

    if ghs_stats:
        headers += ["Объём застройки, тыс. м³", "Площадь коридора, км²"]

    if built_s_stats:
        headers += ["Поверхность застройки, тыс. м²"]

    if overture_stats:
        headers += ["Площадь объектов, тыс. м²"]

    if poi_stats:
        headers += ["Точек POI", "Суммарное value"]

    return headers


def _route_direction_row(
    rd: RouteData,
    di: int,
    d: Any,
    base_row: list[Any],
    *,
    ghs_stats: Mapping[Any, Any] | None,
    ghs_dir_stats: Mapping[Any, Any] | None,
    built_s_stats: Mapping[Any, Any] | None,
    built_s_dir_stats: Mapping[Any, Any] | None,
    overture_stats: Mapping[Any, Any] | None,
    overture_dir_stats: Mapping[Any, Any] | None,
    poi_stats: Mapping[Any, Any] | None,
    poi_dir_stats: Mapping[Any, Any] | None,
) -> list[Any]:
    """Строка одного направления маршрута со статистиками по колонкам."""
    f, t = direction_endpoints(d)

    row = [
        *base_row,
        di + 1,
        d.name or f"{f} → {t}",
        f,
        t,
        d.km,
        fmt_curv_cell(d.curvilinearity),
        len(d.stops),
        rd.url,
    ]

    if ghs_stats:
        st = (ghs_dir_stats or {}).get((rd.route_id, di))
        row += [
            round(st.volume_m3 / 1e3, 1) if st else None,
            round(st.corridor_m2 / 1e6, 3) if st else None,
        ]

    if built_s_stats:
        st = (built_s_dir_stats or {}).get((rd.route_id, di)) or built_s_stats.get(
            rd.route_id
        )
        row += [round(st.surface_m2 / 1e3, 1) if st else None]

    if overture_stats:
        st = (overture_dir_stats or {}).get((rd.route_id, di))
        row += [round(st.total_area_m2 / 1e3, 1) if st else None]

    if poi_stats:
        st = (poi_dir_stats or {}).get((rd.route_id, di))
        row += [
            st.count if st else None,
            round(st.total_value, 1) if st else None,
        ]

    return row


def _style_routes_row(ws: Any, num_fmt: Mapping[int, str], wrap_cols: list[int], link_col: int | None) -> None:
    """Инлайн-оформление последней записанной строки листа «Маршруты»."""
    from openpyxl.styles import Alignment, Font

    r = ws.max_row

    for col, fmt in num_fmt.items():
        ws.cell(row=r, column=col).number_format = fmt

    for col in wrap_cols:
        ws.cell(row=r, column=col).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )

    if link_col:
        cell = ws.cell(row=r, column=link_col)
        if cell.value:
            cell.hyperlink = str(cell.value)
            cell.font = Font(color="0563C1", underline="single")


def _write_routes_sheet(
    wb: Any,
    *,
    city: str,
    kml_routes: Sequence[RouteData],
    ghs_stats: Mapping[Any, Any] | None,
    ghs_dir_stats: Mapping[Any, Any] | None,
    built_s_stats: Mapping[Any, Any] | None,
    built_s_dir_stats: Mapping[Any, Any] | None,
    overture_stats: Mapping[Any, Any] | None,
    overture_dir_stats: Mapping[Any, Any] | None,
    poi_stats: Mapping[Any, Any] | None,
    poi_dir_stats: Mapping[Any, Any] | None,
) -> None:
    """Лист «Маршруты»: по строке на направление с аналитикой."""
    H = _routes_headers(
        ghs_stats=ghs_stats,
        built_s_stats=built_s_stats,
        overture_stats=overture_stats,
        poi_stats=poi_stats,
    )

    link_col = H.index("Ссылка") + 1 if "Ссылка" in H else None

    num_fmt = {
        idx: fmt
        for idx, h in enumerate(H, start=1)
        if (fmt := _ROUTES_NUM_FMT_BY_HEADER.get(h)) is not None
    }

    wrap_cols = [
        idx for idx, h in enumerate(H, start=1) if h in ("Направление", "От", "До")
    ]

    ws = wb.create_sheet("Маршруты")
    ws.append(H)

    # Кэшируем часто используемые значения для маршрутов
    route_label_cache = {}
    route_idea_cache = {}

    for rd in kml_routes:
        route_label_cache[rd.route_id] = _cached_type_label(rd.route_type)
        route_idea_cache[rd.route_id] = rd.is_idea

    for rd in sorted(
        kml_routes,
        key=lambda r: (getattr(r.route_type, "value", r.route_type), r.name),
    ):
        base_row = [
            city,
            route_label_cache[rd.route_id],
            rd.name,
            rd.route_id,
            len(rd.directions),
            "💡 Идея" if route_idea_cache[rd.route_id] else "Реальный",
        ]

        for di, d in enumerate(rd.directions):
            row = _route_direction_row(
                rd,
                di,
                d,
                base_row,
                ghs_stats=ghs_stats,
                ghs_dir_stats=ghs_dir_stats,
                built_s_stats=built_s_stats,
                built_s_dir_stats=built_s_dir_stats,
                overture_stats=overture_stats,
                overture_dir_stats=overture_dir_stats,
                poi_stats=poi_stats,
                poi_dir_stats=poi_dir_stats,
            )

            # Пакетная запись; стили применяются сразу при записи, чтобы
            # не проходить по всем строкам второй раз.
            ws.append(row)

            if num_fmt or wrap_cols or link_col:
                _style_routes_row(ws, num_fmt, wrap_cols, link_col)

    # Ширины колонок задаём фиксированными (ускоряет оформление);
    # если добавлены доп. колонки — дополняем.
    fixed_widths = list(_ROUTES_FIXED_WIDTHS)
    extra = len(H) - len(fixed_widths)

    if extra > 0:
        fixed_widths += [12] * extra

    _finish_sheet(
        ws,
        H,
        num_fmt=num_fmt,
        wrap=set(wrap_cols),
        link_col=link_col,
        fixed_widths=fixed_widths,
        inline_styled=True,
    )


# ═══════════════════════════════════════════════════════════════════════
# ПРОСТЫЕ ЛИСТЫ
# ═══════════════════════════════════════════════════════════════════════
def _write_errors_sheet(wb: Any, city: str, bad: list[RouteData]) -> None:
    """Лист «Ошибки»: маршруты, загрузка которых не удалась."""
    ws = wb.create_sheet("Ошибки")
    H = ["Город", "Тип", "Номер", "ID", "Ошибка", "Ссылка"]
    ws.append(H)

    for r in bad:
        ws.append(
            [city, _cached_type_label(r.route_type), r.name, r.route_id, r.error, r.url]
        )

    _finish_sheet(ws, H, wrap={5}, link_col=6, fixed_widths=[12, 10, 12, 8, 40, 30])


def _write_excluded_sheet(
    wb: Any,
    city: str,
    excluded_stage2: Sequence[tuple[RouteData, str]],
) -> None:
    """Лист «Отсечённые (этап 2)»: маршруты с причиной отсечения."""
    ws = wb.create_sheet("Отсечённые (этап 2)")
    H = ["Город", "Тип", "Номер", "ID", "Причина", "Ссылка"]
    ws.append(H)

    for r, why in excluded_stage2:
        ws.append(
            [city, _cached_type_label(r.route_type), r.name, r.route_id, why, r.url]
        )

    _finish_sheet(ws, H, wrap={5}, link_col=6, fixed_widths=[12, 10, 12, 8, 40, 30])


def _stop_coord(rec: Mapping[str, Any], bbox: tuple[float, float, float, float] | None) -> Any:
    """Исправленные координаты остановки или ``None``."""
    if rec["lat"] is None or rec["lon"] is None:
        return None

    return fix_stop_coord(
        {"latitude": rec["lat"], "longitude": rec["lon"]},
        rec.get("idx", 0),
        bbox,
    )


def _write_unique_stops_sheet(
    wb: Any,
    unique_stops: Mapping[str, Mapping[str, Any]],
    bbox: tuple[float, float, float, float] | None,
) -> None:
    """Лист «Остановки (уникальные)»."""
    ws = wb.create_sheet("Остановки (уникальные)")
    H = ["Название", "ID", "Типы", "Маршрутов", "Широта", "Долгота"]
    ws.append(H)

    for rec in sorted(
        unique_stops.values(),
        key=lambda r: (str(r["name"]).lower(), str(r["id"])),
    ):
        c = _stop_coord(rec, bbox)

        ws.append(
            [
                rec["name"],
                rec["id"] if rec["id"] is not None else "",
                ", ".join(_cached_type_label(t) for t in sorted(rec["types"])),
                len(rec["routes"]),
                c[0] if c else None,
                c[1] if c else None,
            ]
        )

    _finish_sheet(
        ws,
        H,
        num_fmt={5: "0.0000000", 6: "0.0000000"},
        wrap={1, 3},
        fixed_widths=[25, 12, 20, 10, 15, 15],
    )


def _write_stop_volumes_sheet(
    wb: Any,
    unique_stops: Mapping[str, Mapping[str, Any]],
    stop_volumes: Mapping[str, float],
    vol_radius: float | None,
    bbox: tuple[float, float, float, float] | None,
) -> None:
    """Лист «Объём застройки остановок», сортировка по убыванию объёма."""
    ws = wb.create_sheet("Объём застройки остановок")

    rad = vol_radius if vol_radius else 500.0

    H = [
        "Название",
        "ID",
        "Маршрутов",
        "Широта",
        "Долгота",
        f"Объём в радиусе {rad:.0f} м, тыс. м³",
    ]
    ws.append(H)

    rows_data = []

    for key, rec in unique_stops.items():
        c = _stop_coord(rec, bbox)

        rows_data.append(
            [
                rec["name"],
                rec["id"] if rec["id"] is not None else "",
                len(rec["routes"]),
                c[0] if c else None,
                c[1] if c else None,
                round(stop_volumes.get(key, 0.0) / 1e3, 1),
            ]
        )

    rows_data.sort(key=lambda r: -(r[5] or 0))

    for row in rows_data:
        ws.append(list(row))

    _finish_sheet(
        ws,
        H,
        num_fmt={4: "0.0000000", 5: "0.0000000", 6: "#,##0.0"},
        wrap={1},
        fixed_widths=[25, 12, 10, 15, 15, 15],
    )


# ═══════════════════════════════════════════════════════════════════════
# ЛИСТЫ POI
# ═══════════════════════════════════════════════════════════════════════
def _poi_sort_value(r: RouteData, poi_stats: Mapping[Any, Any]) -> float:
    return st.total_value if (st := poi_stats.get(r.route_id)) else 0.0


def _write_poi_sheets(
    wb: Any,
    *,
    city: str,
    ok: list[RouteData],
    poi_stats: Mapping[Any, Any],
    poi_values: Mapping[str, float],
) -> None:
    """Листы «POI вдоль маршрутов» и «Ценность типов POI»."""
    ws = wb.create_sheet("POI вдоль маршрутов")
    H = [
        "Город",
        "Тип",
        "Номер",
        "ID",
        "Точек POI",
        "Суммарное value",
        "Value на км",
    ]

    all_poi_types = sorted(poi_values.keys())

    H.extend(f"{typ}: шт." for typ in all_poi_types)

    H.append("Ссылка")
    ws.append(H)

    link_col_idx = H.index("Ссылка") + 1 if "Ссылка" in H else None

    for rd in sorted(ok, key=lambda r: -_poi_sort_value(r, poi_stats)):
        st = poi_stats.get(rd.route_id)

        val_per_km = (
            round(st.total_value / rd.max_km, 2) if (st and rd.max_km > 0) else None
        )

        row = [
            city,
            _cached_type_label(rd.route_type),
            rd.name,
            rd.route_id,
            st.count if st else 0,
            round(st.total_value, 1) if st else 0.0,
            val_per_km,
        ]

        row.extend(st.by_type.get(typ, 0) if st else 0 for typ in all_poi_types)

        row.append(rd.url)

        ws.append(row)

    num_fmt_dict = {6: "#,##0.0", 7: "0.00"}

    for i in range(8, 8 + len(all_poi_types)):
        num_fmt_dict[i] = "0"

    fixed_widths = [12, 10, 12, 8, 10, 12, 12] + [8] * len(all_poi_types) + [30]

    _finish_sheet(
        ws,
        H,
        num_fmt=num_fmt_dict,
        link_col=link_col_idx,
        fixed_widths=fixed_widths,
    )

    ws2 = wb.create_sheet("Ценность типов POI")
    ws2.append(["Тип POI", "Количество", "Ценность (value)"])

    total_counts: dict[str, int] = {}

    for st in poi_stats.values():
        for typ, cnt in st.by_type.items():
            total_counts[typ] = total_counts.get(typ, 0) + cnt

    for typ, value in sorted(poi_values.items(), key=lambda x: -x[1]):
        ws2.append([typ, total_counts.get(typ, 0), round(value, 2)])

    _finish_sheet(
        ws2,
        ["Тип POI", "Количество", "Ценность (value)"],
        num_fmt={3: "#,##0.00"},
        wrap={1},
        fixed_widths=[25, 12, 15],
    )


# ═══════════════════════════════════════════════════════════════════════
# ЛИСТЫ АНАЛИТИКИ
# ═══════════════════════════════════════════════════════════════════════
def _write_dedup_sheet(
    wb: Any,
    dedup_removed: Sequence[Mapping[str, Any]],
) -> None:
    """Лист «Удаление дублирования»."""
    ws = wb.create_sheet("Удаление дублирования")

    H = [
        "Шаг",
        "Направление (ID)",
        "Тип",
        "Название",
        "Закрывает пар",
        "Партнёры (Kmax)",
        "Причина",
        "Длина, км",
    ]
    ws.append(H)

    for rec in dedup_removed:
        ws.append(
            [
                rec["шаг"],
                rec["маршрут"],
                rec["тип"],
                rec["название"],
                rec["закрывает пар"],
                rec["партнёры (Kmax)"],
                rec["причина"],
                rec["длина, км"],
            ]
        )

    _finish_sheet(
        ws,
        H,
        num_fmt={
            8: "0.00",
            9: "#,##0.0",
        },
        wrap={6, 7},
        fixed_widths=[
            8,  # Шаг
            18,  # Направление (ID)
            10,  # Тип
            20,  # Название
            15,  # Закрывает пар
            20,  # Партнёры (Kmax)
            30,  # Причина
            10,  # Длина, км
        ],
    )


def _write_generated_sheet(
    wb: Any,
    generated_routes: Sequence[Mapping[str, Any]],
) -> None:
    """Лист «Сгенерированные маршруты»."""
    ws = wb.create_sheet("Сгенерированные маршруты")

    H = [
        "№",
        "От",
        "До",
        "Длина, км",
        "Остановок",
        "Прямолинейность",
        "Объём коридора, тыс. м³",
        "Объём/км, тыс. м³",
        "Статус",
        "Трасса (остановки)",
    ]
    ws.append(H)

    for g in generated_routes:
        ws.append(
            [
                g["n"],
                g["from"],
                g["to"],
                round(g["length_km"], 2),
                g["stops"],
                fmt_curv_cell(g["curvilinearity"]),
                round(g["volume"] / 1e3, 1),
                round(g["volume"] / 1e3 / g["length_km"], 1)
                if g["length_km"] > 0
                else None,
                g["status"],
                g["chain"],
            ]
        )

    _finish_sheet(
        ws,
        H,
        num_fmt={4: "0.00", 6: "0.00", 7: "#,##0.0", 8: "#,##0.0"},
        wrap={2, 3, 10},
        fixed_widths=[6, 20, 20, 10, 8, 10, 12, 12, 10, 40],
    )


def _write_heatmap_sheet(wb: Any, heatmap: Mapping[str, Any]) -> None:
    """Лист «Тепловая карта»."""
    ws = wb.create_sheet("Тепловая карта")
    H = ["№", "Широта центра", "Долгота центра", "Значение", "Доля от макс, %"]
    ws.append(H)

    for n, cell in enumerate(heatmap["cells"], 1):
        ws.append(
            [
                n,
                round((cell["lat0"] + cell["lat1"]) / 2, 6),
                round((cell["lon0"] + cell["lon1"]) / 2, 6),
                round(cell["value"], 1),
                round(cell["t"] * 100, 1),
            ]
        )

    _finish_sheet(
        ws,
        H,
        num_fmt={2: "0.000000", 3: "0.000000", 5: "0.0"},
        fixed_widths=[6, 15, 15, 12, 12],
    )


def _write_all_stops_sheet(
    wb: Any,
    *,
    city: str,
    ok: list[RouteData],
    bbox: tuple[float, float, float, float] | None,
) -> None:
    """Лист «Остановки»: все остановки всех успешных маршрутов."""
    ws = wb.create_sheet("Остановки")
    H = [
        "Город",
        "Тип",
        "Номер",
        "ID",
        "№ напр.",
        "№ ост.",
        "Название",
        "Широта",
        "Долгота",
    ]
    ws.append(H)

    for rd in ok:
        type_lbl = _cached_type_label(rd.route_type)

        for di, d in enumerate(rd.directions):
            for si, s in enumerate(d.stops):
                if not isinstance(s, dict):
                    continue

                c = fix_stop_coord(s, si, bbox)

                ws.append(
                    [
                        city,
                        type_lbl,
                        rd.name,
                        rd.route_id,
                        di + 1,
                        si + 1,
                        str(s.get("name", "")),
                        c[0] if c else None,
                        c[1] if c else None,
                    ]
                )

    _finish_sheet(
        ws,
        H,
        num_fmt={8: "0.0000000", 9: "0.0000000"},
        wrap={7},
        fixed_widths=[12, 10, 12, 8, 8, 8, 25, 15, 15],
    )


# ═══════════════════════════════════════════════════════════════════════
# ТОЧКА ВХОДА
# ═══════════════════════════════════════════════════════════════════════
def build_xlsx(
    routes: Sequence[RouteData],
    city: str,
    city_title: str,
    path: str | Path,
    curv_limit: float = 0.0,
    minlen_limit: float = 0.0,
    radius_limit: float = 0.0,
    cut_curv: int = 0,
    cut_len: int = 0,
    cut_radius: int = 0,
    skipped_inactive: int = 0,
    kml_routes: Sequence[RouteData] | None = None,
    unique_stops: Mapping[str, Mapping[str, Any]] | None = None,
    excluded_stage2: Sequence[tuple[RouteData, str]] | None = None,
    heatmap: Mapping[str, Any] | None = None,
    include_stops: bool = False,
    bbox: tuple[float, float, float, float] | None = None,
    ghs_stats: Mapping[Any, Any] | None = None,
    ghs_meta: Mapping[str, Any] | None = None,
    ghs_dir_stats: Mapping[Any, Any] | None = None,
    poi_stats: Mapping[Any, Any] | None = None,
    poi_values: Mapping[str, float] | None = None,
    poi_buffer_m: float = 0.0,
    poi_dir_stats: Mapping[Any, Any] | None = None,
    dedup_removed: Sequence[Mapping[str, Any]] | None = None,
    dedup_analysis: Mapping[str, Any] | None = None,
    net_metrics: Mapping[str, Any] | None = None,
    generated_routes: Sequence[Mapping[str, Any]] | None = None,
    built_s_stats: Mapping[Any, Any] | None = None,
    built_s_meta: Mapping[str, Any] | None = None,
    built_s_dir_stats: Mapping[Any, Any] | None = None,
    overture_stats: Mapping[Any, Any] | None = None,
    overture_meta: Mapping[str, Any] | None = None,
    overture_dir_stats: Mapping[Any, Any] | None = None,
    stop_volumes: Mapping[str, float] | None = None,
    vol_radius: float | None = None,
) -> str | Path | None:
    """Создаёт XLSX-отчёт по сети маршрутов и возвращает путь к файлу.

    Параметры аналитики (`ghs_*`, `poi_*`, `dedup_*`, `overture_*`) принимаются
    как mapping-объекты, сформированные соответствующими вычислительными модулями.

    Неактивные/ошибочные маршруты используются только для соответствующих листов,
    а `openpyxl` остаётся optional dependency и проверяется при вызове.
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        logger.warning("XLSX пропущен: pip install openpyxl")
        return None

    ok = [r for r in routes if not r.error and r.directions]
    bad = [r for r in routes if r.error]
    ideas = [r for r in routes if r.is_idea and not r.error]

    wb = Workbook()

    _write_summary_sheet(
        wb,
        _summary_rows(
            city=city,
            city_title=city_title,
            routes=routes,
            ok=ok,
            bad=bad,
            ideas=ideas,
            skipped_inactive=skipped_inactive,
            curv_limit=curv_limit,
            minlen_limit=minlen_limit,
            radius_limit=radius_limit,
            cut_curv=cut_curv,
            cut_len=cut_len,
            cut_radius=cut_radius,
            dedup_removed=dedup_removed,
            dedup_analysis=dedup_analysis,
            net_metrics=net_metrics,
            net_density=_network_density(net_metrics, bbox),
            kml_routes=kml_routes,
            unique_stops=unique_stops,
            generated_routes=generated_routes,
            bbox=bbox,
            ghs_stats=ghs_stats,
            ghs_meta=ghs_meta,
            poi_buffer_m=poi_buffer_m,
            poi_stats=poi_stats,
            built_s_stats=built_s_stats,
            built_s_meta=built_s_meta,
            overture_stats=overture_stats,
            overture_meta=overture_meta,
        ),
    )

    if kml_routes:
        _write_routes_sheet(
            wb,
            city=city,
            kml_routes=kml_routes,
            ghs_stats=ghs_stats,
            ghs_dir_stats=ghs_dir_stats,
            built_s_stats=built_s_stats,
            built_s_dir_stats=built_s_dir_stats,
            overture_stats=overture_stats,
            overture_dir_stats=overture_dir_stats,
            poi_stats=poi_stats,
            poi_dir_stats=poi_dir_stats,
        )

    _write_errors_sheet(wb, city, bad)

    if excluded_stage2:
        _write_excluded_sheet(wb, city, excluded_stage2)

    if unique_stops:
        _write_unique_stops_sheet(wb, unique_stops, bbox)

    if stop_volumes and unique_stops:
        _write_stop_volumes_sheet(wb, unique_stops, stop_volumes, vol_radius, bbox)

    if poi_stats and poi_values:
        _write_poi_sheets(wb, city=city, ok=ok, poi_stats=poi_stats, poi_values=poi_values)

    if dedup_removed:
        _write_dedup_sheet(wb, dedup_removed)

    if generated_routes:
        _write_generated_sheet(wb, generated_routes)

    if heatmap:
        _write_heatmap_sheet(wb, heatmap)

    if include_stops:
        _write_all_stops_sheet(wb, city=city, ok=ok, bbox=bbox)

    try:
        wb.save(path)
    except (PermissionError, OSError) as e:
        logger.warning("Не удалось сохранить %s: %s", path, e)
        return None

    return path
