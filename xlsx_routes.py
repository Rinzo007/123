"""Лист «Маршруты» XLSX-экспорта."""

from __future__ import annotations

from collections.abc import Mapping, Sequence
from typing import Any

from .models import RouteData
from .support import direction_endpoints, fmt_curv_cell, type_label
from .xlsx_helpers import cached_type_label, finish_sheet

_ROUTES_BASE_HEADERS = [
    "Город", "Тип", "Номер", "ID", "Направлений", "Источник", "№ напр.",
    "Направление", "От", "До", "Длина, км", "Коэфф. непрямолинейности",
    "Остановок", "Ссылка",
]

_ROUTES_NUM_FMT_BY_HEADER = {
    "Длина, км": "0.0",
    "Коэфф. непрямолинейности": "0.00",
    "Остановок": "0",
    "Объём застройки, тыс. м³": "#,##0.0",
    "Площадь коридора, км²": "0.000",
    "Поверхность застройки, тыс. м²": "#,##0.0",
    "Площадь объектов, тыс. м²": "#,##0.0",
    "Точек POI": "0",
    "Суммарное value": "#,##0.0",
}

_ROUTES_FIXED_WIDTHS = [12, 10, 12, 8, 10, 12, 8, 20, 15, 15, 10, 10, 10, 30]


def routes_headers(*, ghs_stats=None, built_s_stats=None, overture_stats=None, poi_stats=None) -> list[str]:
    headers = list(_ROUTES_BASE_HEADERS)
    if ghs_stats:
        headers += ["Объём застройки, тыс. м³", "Площадь коридора, км²"]
    if built_s_stats:
        headers += ["Поверхность застройки, тыс. м²"]
    if overture_stats:
        headers += ["Площадь объектов, тыс. м²"]
    if poi_stats:
        headers += ["Точек POI", "Суммарное value"]
    return headers


def route_direction_row(rd: RouteData, di: int, direction: Any, base_row: list[Any], *,
                        ghs_stats=None, ghs_dir_stats=None, built_s_stats=None,
                        built_s_dir_stats=None, overture_stats=None,
                        overture_dir_stats=None, poi_stats=None, poi_dir_stats=None) -> list[Any]:
    first, last = direction_endpoints(direction)
    row = [*base_row, di + 1, direction.name or f"{first} → {last}", first, last,
           direction.km, fmt_curv_cell(direction.curvilinearity), len(direction.stops), rd.url]
    if ghs_stats:
        st = (ghs_dir_stats or {}).get((rd.route_id, di))
        row += [round(st.volume_m3 / 1e3, 1) if st else None,
                round(st.corridor_m2 / 1e6, 3) if st else None]
    if built_s_stats:
        st = (built_s_dir_stats or {}).get((rd.route_id, di)) or built_s_stats.get(rd.route_id)
        row += [round(st.surface_m2 / 1e3, 1) if st else None]
    if overture_stats:
        st = (overture_dir_stats or {}).get((rd.route_id, di))
        row += [round(st.total_area_m2 / 1e3, 1) if st else None]
    if poi_stats:
        st = (poi_dir_stats or {}).get((rd.route_id, di))
        row += [st.count if st else None, round(st.total_value, 1) if st else None]
    return row


def _style_row(ws: Any, num_fmt: Mapping[int, str], wrap_cols: list[int], link_col: int | None) -> None:
    from openpyxl.styles import Alignment, Font
    row = ws.max_row
    for col, fmt in num_fmt.items():
        ws.cell(row=row, column=col).number_format = fmt
    for col in wrap_cols:
        ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
    if link_col:
        cell = ws.cell(row=row, column=link_col)
        if cell.value:
            cell.hyperlink = str(cell.value)
            cell.font = Font(color="0563C1", underline="single")


def write_routes_sheet(wb: Any, *, city: str, kml_routes: Sequence[RouteData],
                       ghs_stats=None, ghs_dir_stats=None, built_s_stats=None,
                       built_s_dir_stats=None, overture_stats=None,
                       overture_dir_stats=None, poi_stats=None, poi_dir_stats=None) -> None:
    headers = routes_headers(ghs_stats=ghs_stats, built_s_stats=built_s_stats,
                             overture_stats=overture_stats, poi_stats=poi_stats)
    link_col = headers.index("Ссылка") + 1 if "Ссылка" in headers else None
    num_fmt = {i: fmt for i, h in enumerate(headers, 1) if (fmt := _ROUTES_NUM_FMT_BY_HEADER.get(h))}
    wrap_cols = [i for i, h in enumerate(headers, 1) if h in ("Направление", "От", "До")]
    ws = wb.create_sheet("Маршруты")
    ws.append(headers)

    route_label_cache: dict[int, str] = {}
    route_idea_cache: dict[int, bool] = {}
    for route in kml_routes:
        route_label_cache[route.route_id] = cached_type_label(route.route_type)
        route_idea_cache[route.route_id] = route.is_idea

    for route in sorted(kml_routes, key=lambda r: (getattr(r.route_type, "value", r.route_type), r.name)):
        base_row = [city, route_label_cache[route.route_id], route.name, route.route_id,
                    len(route.directions), "💡 Идея" if route_idea_cache[route.route_id] else "Реальный"]
        for di, direction in enumerate(route.directions):
            ws.append(route_direction_row(route, di, direction, base_row,
                                           ghs_stats=ghs_stats, ghs_dir_stats=ghs_dir_stats,
                                           built_s_stats=built_s_stats, built_s_dir_stats=built_s_dir_stats,
                                           overture_stats=overture_stats, overture_dir_stats=overture_dir_stats,
                                           poi_stats=poi_stats, poi_dir_stats=poi_dir_stats))
            if num_fmt or wrap_cols or link_col:
                _style_row(ws, num_fmt, wrap_cols, link_col)

    widths = list(_ROUTES_FIXED_WIDTHS)
    if len(headers) > len(widths):
        widths += [12] * (len(headers) - len(widths))
    finish_sheet(ws, headers, num_fmt=num_fmt, wrap=set(wrap_cols), link_col=link_col,
                 fixed_widths=widths, inline_styled=True)
